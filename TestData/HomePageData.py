import openpyxl


class HomePageData:
    test_HomePage_Data = [{"firstname":"veda","email":"KM","gender":"Female"},{"firstname":"megha","email":"KM", "gender":"Male"}]

    @staticmethod  #declaring this method as static inorder to access this method using class name instead of creating object of that class
    #def getTestData(self,test_case_name):  #remove self keyword as this is required fro only non static methods
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\vinay\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # for j in range(1,sheet.max_column+1):
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        #print(Dict)
        return [Dict]